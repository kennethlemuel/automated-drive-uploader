#!/usr/bin/env python3
import argparse
import mimetypes
import warnings
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

SCOPES = ["https://www.googleapis.com/auth/drive.file"]
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
LOCAL_IMAGE_HEADERS = {"ad screenshot"}


def image_row(image):
    return image.anchor._from.row + 1


def image_col(image):
    return image.anchor._from.col + 1


def image_ext(image):
    path = Path(getattr(image, "path", "image.png"))
    return path.suffix or ".png"


def image_bytes(image):
    data = image._data()
    image.ref = BytesIO(data)
    return data


def resolve_image_path(line, workbook_dir, image_root=None):
    raw = Path(line)
    if raw.is_absolute():
        return raw

    candidates = [workbook_dir / raw]
    if image_root:
        image_root = Path(image_root)
        candidates.append(image_root / raw)
        parts = raw.parts
        if parts and parts[0] == image_root.name:
            candidates.append(image_root.joinpath(*parts[1:]))
        candidates.append(image_root.parent / raw)

    for path in candidates:
        if path.exists():
            return path.resolve()
    return candidates[0].resolve()


def cell_image_paths(value, workbook_dir, image_root=None):
    if not isinstance(value, str):
        return []
    paths = []
    for line in value.splitlines():
        line = line.strip()
        if not line or line.startswith(("http://", "https://")):
            continue
        path = resolve_image_path(line, workbook_dir, image_root)
        if path.suffix.lower() in IMAGE_EXTS:
            paths.append(path)
    return paths


def local_image_columns(sheet):
    columns = set()
    for cell in sheet[1]:
        value = str(cell.value or "").strip().lower()
        if value in LOCAL_IMAGE_HEADERS:
            columns.add(cell.column)
    return columns


def write_link_cell(cell, links):
    if len(links) == 1:
        url, label = links[0]
        cell.value = label
        cell.hyperlink = url
        cell.style = "Hyperlink"
        return
    cell.value = "\n".join(f"{label}: {url}" for url, label in links)


def count_uploads(workbook, input_path, image_root=None):
    total = 0
    for sheet in workbook.worksheets:
        image_columns = local_image_columns(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                if image_columns and cell.column not in image_columns:
                    continue
                total += sum(path.exists() for path in cell_image_paths(cell.value, input_path.parent, image_root))
        total += len(getattr(sheet, "_images", []))
    return total


def auth_drive(credentials_path, token_path):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    token = Path(token_path)
    if token.exists():
        creds = Credentials.from_authorized_user_file(token, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        token.write_text(creds.to_json())
    return build("drive", "v3", credentials=creds)


def upload_image(service, data, name, folder_id=None, public=True):
    from googleapiclient.http import MediaInMemoryUpload

    mime_type = mimetypes.guess_type(name)[0] or "application/octet-stream"
    body = {"name": name}
    if folder_id:
        body["parents"] = [folder_id]
    created = service.files().create(
        body=body,
        media_body=MediaInMemoryUpload(data, mimetype=mime_type, resumable=False),
        fields="id",
    ).execute()
    file_id = created["id"]
    if public:
        service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
            fields="id",
        ).execute()
    return f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"


def upload_workbook(service, path, folder_id=None, public=True, as_sheets=True):
    from googleapiclient.http import MediaFileUpload

    path = Path(path)
    body = {"name": path.stem if as_sheets else path.name}
    if as_sheets:
        body["mimeType"] = "application/vnd.google-apps.spreadsheet"
    if folder_id:
        body["parents"] = [folder_id]
    created = service.files().create(
        body=body,
        media_body=MediaFileUpload(
            path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False,
        ),
        fields="id,webViewLink",
    ).execute()
    if public:
        service.permissions().create(
            fileId=created["id"],
            body={"type": "anyone", "role": "reader"},
            fields="id",
        ).execute()
    return created.get("webViewLink") or f"https://drive.google.com/file/d/{created['id']}/view"


def convert(input_path, output_path, upload, image_root=None, progress=None):
    input_path = Path(input_path)
    uploaded_count = 0
    missing_paths = []
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="DrawingML support is incomplete.*", module="openpyxl.reader.drawings")
        workbook = load_workbook(input_path)
    total_uploads = count_uploads(workbook, input_path, image_root)
    for sheet in workbook.worksheets:
        images = list(getattr(sheet, "_images", []))
        links_by_cell = {}
        pending = []
        image_columns = local_image_columns(sheet)

        for row in sheet.iter_rows():
            for cell in row:
                if image_columns and cell.column not in image_columns:
                    continue
                for path in cell_image_paths(cell.value, input_path.parent, image_root):
                    if path.exists():
                        pending.append((cell.row, cell.column if image_columns else None, path.name, lambda p=path: p.read_bytes()))
                    else:
                        missing_paths.append(str(path))

        if not images and not pending:
            continue

        link_col = min(image_columns) if image_columns else (max(sheet.max_column, *(image_col(image) for image in images)) + 1 if images else sheet.max_column + 1)
        if not image_columns:
            sheet.cell(row=1, column=link_col).value = "Image Link"

        for index, image in enumerate(images, start=1):
            row = image_row(image)
            name = f"{Path(input_path).stem}-{sheet.title}-r{row}-{index}{image_ext(image)}"
            pending.append((row, link_col, name, lambda img=image: image_bytes(img)))

        for row, column, name, read_data in pending:
            links_by_cell.setdefault((row, column or link_col), []).append((upload(read_data(), name), name))
            uploaded_count += 1
            if progress:
                progress(uploaded_count, total_uploads, name)

        for (row, column), links in links_by_cell.items():
            write_link_cell(sheet.cell(row=row, column=column), links)

    workbook.save(output_path)
    return uploaded_count, missing_paths


def main():
    parser = argparse.ArgumentParser(description="Upload embedded XLSX images to Google Drive and write links to a copy.")
    parser.add_argument("input", help="Path to the .xlsx file")
    parser.add_argument("-o", "--output", help="Output .xlsx path")
    parser.add_argument("--credentials", help="Google OAuth desktop-client JSON")
    parser.add_argument("--token", default=".google-drive-token.json", help="OAuth token cache path")
    parser.add_argument("--folder-id", help="Google Drive folder ID for uploads")
    parser.add_argument("--image-root", help="Folder containing screenshot files referenced by the workbook")
    parser.add_argument("--upload-workbook", action="store_true", help="Upload the output workbook to Google Sheets")
    parser.add_argument("--keep-xlsx", action="store_true", help="Upload the output workbook as .xlsx instead of converting to Google Sheets")
    parser.add_argument("--private", action="store_true", help="Do not create public link-view permissions")
    parser.add_argument("--dry-run", action="store_true", help="Write fake links without uploading")
    args = parser.parse_args()

    input_path = Path(args.input)
    if input_path.suffix.lower() != ".xlsx":
        raise SystemExit("Only .xlsx files are supported.")

    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}-with-links.xlsx")

    if args.dry_run:
        upload = lambda _data, name: f"dry-run://{name}"
    else:
        if not args.credentials:
            raise SystemExit("--credentials is required unless --dry-run is used.")
        service = auth_drive(args.credentials, args.token)
        upload = lambda data, name: upload_image(service, data, name, args.folder_id, not args.private)

    uploaded_count, missing_paths = convert(input_path, output_path, upload, args.image_root)
    print(output_path)
    print(f"uploaded={uploaded_count}")
    if args.upload_workbook and not args.dry_run:
        print(upload_workbook(service, output_path, args.folder_id, not args.private, not args.keep_xlsx))
    if missing_paths:
        print(f"missing_local_files={len(missing_paths)}")
        for path in missing_paths[:10]:
            print(path)
    if uploaded_count == 0:
        print("No embedded images or existing local image paths were found.")


if __name__ == "__main__":
    main()
