from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from excel_drive_links import convert


def test_embedded_image_gets_row_link():
    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        png = tmp / "sample.png"
        source = tmp / "source.xlsx"
        output = tmp / "output.xlsx"

        PILImage.new("RGB", (8, 8), "red").save(png)

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["A2"] = "Product"
        sheet.add_image(Image(png), "B2")
        workbook.save(source)

        convert(source, output, lambda _data, name: f"dry-run://{name}")

        original = load_workbook(source)
        result = load_workbook(output)

        assert original.active["C1"].value is None
        assert result.active["C1"].value == "Image Link"
        assert result.active["C2"].value.startswith("source-Sheet-r2-1")
        assert result.active["C2"].hyperlink.target.startswith("dry-run://source-Sheet-r2-1")


def test_local_image_path_gets_row_link():
    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        screenshot_dir = tmp / "Analytics Audit Screenshots"
        screenshot_dir.mkdir()
        png = screenshot_dir / "click.png"
        source = tmp / "source.xlsx"
        output = tmp / "output.xlsx"

        PILImage.new("RGB", (8, 8), "blue").save(png)

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["A2"] = "Product"
        sheet["B2"] = "Analytics Audit Screenshots/click.png"
        workbook.save(source)

        uploaded, missing = convert(source, output, lambda _data, name: f"dry-run://{name}")
        result = load_workbook(output)

        assert uploaded == 1
        assert missing == []
        assert result.active["C1"].value == "Image Link"
        assert result.active["C2"].value == "click.png"
        assert result.active["C2"].hyperlink.target == "dry-run://click.png"


def test_local_image_path_can_use_selected_screenshot_folder():
    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        screenshot_dir = tmp / "Elsewhere" / "Analytics Audit Screenshots"
        screenshot_dir.mkdir(parents=True)
        png = screenshot_dir / "click.png"
        source = tmp / "source.xlsx"
        output = tmp / "output.xlsx"

        PILImage.new("RGB", (8, 8), "green").save(png)

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["A2"] = "Product"
        sheet["B2"] = "Analytics Audit Screenshots/click.png"
        workbook.save(source)

        uploaded, missing = convert(source, output, lambda _data, name: f"dry-run://{name}", screenshot_dir)
        result = load_workbook(output)

        assert uploaded == 1
        assert missing == []
        assert result.active["C2"].value == "click.png"
        assert result.active["C2"].hyperlink.target == "dry-run://click.png"


def test_nearby_screenshots_are_ignored_when_ad_screenshot_column_exists():
    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        screenshot_dir = tmp / "Analytics Audit Screenshots"
        screenshot_dir.mkdir()
        ad_png = screenshot_dir / "ad.png"
        nearby_png = screenshot_dir / "nearby.png"
        source = tmp / "source.xlsx"
        output = tmp / "output.xlsx"

        PILImage.new("RGB", (8, 8), "green").save(ad_png)
        PILImage.new("RGB", (8, 8), "yellow").save(nearby_png)

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Name"
        sheet["B1"] = "AD Screenshot"
        sheet["C1"] = "Nearby Screenshots"
        sheet["A2"] = "Product"
        sheet["B2"] = "Analytics Audit Screenshots/ad.png"
        sheet["C2"] = "Analytics Audit Screenshots/nearby.png"
        workbook.save(source)

        uploaded, missing = convert(source, output, lambda _data, name: f"dry-run://{name}")
        result = load_workbook(output)

        assert uploaded == 1
        assert missing == []
        assert result.active["D2"].value == "ad.png"
        assert result.active["D2"].hyperlink.target == "dry-run://ad.png"


def test_progress_reports_global_percentage():
    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        screenshot_dir = tmp / "Analytics Audit Screenshots"
        screenshot_dir.mkdir()
        first = screenshot_dir / "first.png"
        second = screenshot_dir / "second.png"
        source = tmp / "source.xlsx"
        output = tmp / "output.xlsx"

        PILImage.new("RGB", (8, 8), "red").save(first)
        PILImage.new("RGB", (8, 8), "blue").save(second)

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "AD Screenshot"
        sheet["A2"] = "Analytics Audit Screenshots/first.png"
        sheet["A3"] = "Analytics Audit Screenshots/second.png"
        workbook.save(source)

        seen = []
        convert(source, output, lambda _data, name: f"dry-run://{name}", progress=lambda done, total, _name: seen.append((done, total)))

        assert seen == [(1, 2), (2, 2)]


if __name__ == "__main__":
    test_embedded_image_gets_row_link()
    test_local_image_path_gets_row_link()
    test_local_image_path_can_use_selected_screenshot_folder()
    test_nearby_screenshots_are_ignored_when_ad_screenshot_column_exists()
    test_progress_reports_global_percentage()
